"""
Finance Hours Reconciliation
Runs every Wednesday — compares vacation leave with Allocation sheet and updates discrepancies.
"""

import os
import json
import base64
import tempfile
import requests
from datetime import datetime, timedelta, date
from google.oauth2 import service_account
from googleapiclient.discovery import build

# ── CONFIG ────────────────────────────────────────────────────────────────────

ALLOCATION_SHEET_ID  = "1LIbxO1gpQVX-QDRQXUfS497eX0aFT8t1YtxYCs4lx9k"
ALLOCATION_TAB       = "Allocation- changes PROHIBITED!!!"
ALLOCATION_SHEET_NUM = 2120159953
GITHUB_REPO          = "samciandt/Finance"
GITHUB_FILE          = "2026 Biz hours.xlsx"
TARGET_YEAR          = 2026
VALID_STATUS         = {"approved", "requested", "intention"}

TEAMS = [
    {
        "name":         "Shopping_Tools_Transactions",
        "alloc_key":    "Audi of America.Shopping_Tools_Transactions-High_Value",
        "leave_sheet":  "1JgQG1vZ6k0G9LxiwBlTkBVZnziqlnYt0W3E069dgmSE",
        "leave_tab":    "2026_LEAVE",
        # Explicit roster: leave sheet has no Location column, so auto-detect fails
        "roster": [
            {"name": "Samla Gorza Borges",                   "login": "samla",           "loc": "CPS - Campinas"},
            {"name": "Larissa Virginia Dos Santos Pinheiro", "login": "lpinheiro",        "loc": "CPS - Campinas"},
            {"name": "Bryan Tovar",                          "login": "bryan.tovar",      "loc": "COL - Medellin"},
            {"name": "James Gutierrez",                      "login": "james.gutierrez",  "loc": "COL - Medellin"},
            {"name": "Tim Pepper",                           "login": "tim.pepper",       "loc": "US"},
        ],
    },
    {
        "name":         "Charging",
        "alloc_key":    "Audi of America.Charging-High_Value",
        "leave_sheet":  "11WN88HrRW_GAyrAMY-3p3PHL2im0yu9BUjDX2eUFnCs",
        "leave_tab":    "2026_LEAVE",
        # christinan is a shared resource (50% Charging / 50% eCommerce-Portfolio).
        # Her leave data lives here (Charging), so she must be in this roster so that
        # reconcile_and_update also updates her eCommerce-Portfolio row via the
        # shared-resource bypass (any login in SHARED_RESOURCES is updated in ALL
        # allocation rows where it appears, not just this team's rows).
    },
    {
        "name":         "MyAudi",
        "alloc_key":    "Audi of America.myAudi-High_Value",
        "leave_sheet":  "1rc5DelFVBhNNIboxzD52UyZ56ZMeAmICCWlTbjhBCJU",
        "leave_tab":    "2026_LEAVE",
    },
    {
        "name":         "Digital_Products",
        "alloc_key":    "Audi of America.eCommerce-Portfolio",
        "leave_sheet":  "1WqfsdAGI3YSBW6KrxaOOqEHL1gL20-CW9Us7cPl5jPE",
        "leave_tab":    "2026_LEAVE",
        # Explicit roster: leave sheet name→login mapping
        # christinan excluded here — she is handled via Charging team (shared resource bypass)
        "roster": [
            {"name": "Diego Fernando Benavides Ariza", "login": "diego.ariza",      "loc": "COL - Medellin"},
            {"name": "Enaile Caldas Rebello",          "login": "enaile.rebello",   "loc": "SP - Sao Paulo"},
            {"name": "Gabriela Lozano Ospina",         "login": "gabriela.lozano",  "loc": "LIS - Lisbon"},
            {"name": "Gabriell Silveira Santos",       "login": "gabriellsantos",   "loc": "CPS - Campinas"},
            {"name": "Miguel Figueroa",                "login": "miguel.figueroa",  "loc": "COL - Medellin"},
            {"name": "Pedro Henrique Botecchi",        "login": "pbotecchi",        "loc": "CPS - Campinas"},
            {"name": "Sarah Bizal",                    "login": "sarah.bizal",      "loc": "US"},
            {"name": "Stephen Boyton",                 "login": "stephen.boynton",  "loc": "US"},
            {"name": "Tara Gass",                      "login": "tara.gass",        "loc": "US"},
        ],
    },
    {
        "name":         "Prod_Support",
        "alloc_key":    "Audi of America.Prod_Support-High_Value",
        "leave_sheet":  "1gFe1uK21Abr7JiHdxLA_P-vSttPUbE85SztaGmWb4kY",
        "leave_tab":    "2026_LEAVE",
    },
    {
        "name":         "Enablement",
        "alloc_key":    "Audi of America.Leadership-High_Value",
        "leave_sheet":  "1DZKwHngg7AavyBQ_8SLQMso8Lc4Jpgxf8Fy7R9ReuIo",
        "leave_tab":    "2026_LEAVE",
        # TODO: leave sheet names (Isabel Graziela, Gustavo Alves Vasconcelos,
        # Mathews Palumbo, Daniel Medina Cossio, Thiago De Macedo Bartoleti)
        # do not map to allocation logins (ericav, bruno, jpedretti, hannah.lee,
        # alex.gass, ben). Add explicit roster once login mapping is confirmed.
        "skip": True,
    },
]

# Shared resources: login → factor per team (applied globally across all teams).
# reconcile_and_update updates ALL allocation rows for these logins, not just the
# current team's rows — so leave data only needs to exist in one team's leave sheet.
SHARED_RESOURCES = {
    "christinan": 0.5,   # 50% Charging / 50% eCommerce-Portfolio
    "tara.gass":  0.25,  # 25% per team
}

LOC_MAP = {
    "CPS": "CPS - Campinas", "SP": "SP - Sao Paulo",
    "BH": "BH - Belo Horizonte", "US": "US",
    "CO": "COL - Medellin", "COL": "COL - Medellin",
    "CA": "US", "SEA": "SEA - Seattle", "MNL": "MNL - Manila",
    "LIS": "LIS - Lisbon", "TOR": "TOR - Toronto",
    "ESP": "LIS - Lisbon",  # Spain — using Lisbon business calendar as proxy
}

MONTH_COL = {1:23,2:24,3:25,4:26,5:27,6:28,7:29,8:30,9:31,10:32,11:33,12:34}

ORANGE = {"red": 1.0, "green": round(109/255, 6), "blue": round(1/255, 6)}


# ── AUTH ──────────────────────────────────────────────────────────────────────

def get_sheets_service():
    key_json = os.environ.get("GOOGLE_SERVICE_ACCOUNT_JSON", "").strip()
    oauth_token = os.environ.get("GOOGLE_OAUTH_TOKEN", "").strip()

    if key_json:
        try:
            creds = service_account.Credentials.from_service_account_info(
                json.loads(key_json),
                scopes=["https://www.googleapis.com/auth/spreadsheets"]
            )
            return build("sheets", "v4", credentials=creds, cache_discovery=False)
        except Exception as e:
            print(f"[WARN] Service account auth failed: {e}")

    if oauth_token:
        from google.oauth2.credentials import Credentials as OAuthCreds
        creds = OAuthCreds(token=oauth_token)
        return build("sheets", "v4", credentials=creds, cache_discovery=False)

    raise RuntimeError(
        "No valid Google credentials found.\n"
        "Set GOOGLE_SERVICE_ACCOUNT_JSON (service account key JSON) — permanent.\n"
        "Or set GOOGLE_OAUTH_TOKEN for temporary testing (expires in ~1h)."
    )


def get_github_token():
    return os.environ["GITHUB_PAT"]


def get_gchat_webhook():
    return os.environ["GCHAT_WEBHOOK"]


# ── BIZ DAYS ─────────────────────────────────────────────────────────────────

def download_biz_xlsx(github_token):
    import openpyxl
    url = f"https://api.github.com/repos/{GITHUB_REPO}/contents/{GITHUB_FILE.replace(' ', '%20')}"
    resp = requests.get(url, headers={"Authorization": f"Bearer {github_token}"})
    resp.raise_for_status()
    content = resp.json()["content"].replace("\n", "").replace("\r", "").replace(" ", "")
    data = base64.b64decode(content)
    tmp = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
    tmp.write(data)
    tmp.close()
    return openpyxl.load_workbook(tmp.name, data_only=True)


def parse_biz_data(wb):
    biz_days = {}
    holidays = {}
    for loc in wb.sheetnames:
        ws = wb[loc]
        hols, in_summary = [], False
        biz_days[loc] = {}
        month_map = {"Jan":1,"Feb":2,"Mar":3,"Apr":4,"May":5,"Jun":6,
                     "Jul":7,"Aug":8,"Sep":9,"Oct":10,"Nov":11,"Dec":12}
        for row in ws.iter_rows(values_only=True):
            if row[0] == "Monthly Business Hours Summary":
                in_summary = True
                continue
            if in_summary and row[0] in month_map:
                biz_days[loc][month_map[row[0]]] = row[3]
            if not in_summary and isinstance(row[0], str) and "/" in row[0]:
                try:
                    d = datetime.strptime(row[0], "%m/%d/%Y")
                    hols.append(d.strftime("%Y-%m-%d"))
                except:
                    pass
        holidays[loc] = hols
    return biz_days, holidays


# ── LEAVE CALCULATION ────────────────────────────────────────────────────────

def count_leave_days(start_str, end_str, month, hols):
    try:
        start = datetime.strptime(start_str.strip(), "%m/%d/%Y").date()
        end   = datetime.strptime(end_str.strip(),   "%m/%d/%Y").date()
    except:
        return 0
    count, cursor = 0, start
    while cursor <= end:
        if (cursor.year == TARGET_YEAR
                and cursor.month == month
                and cursor.weekday() < 5
                and cursor.strftime("%Y-%m-%d") not in hols):
            count += 1
        cursor += timedelta(days=1)
    return count


def calc_person_hours(name, loc, leave_rows, biz_days, holidays, months, factor=1.0):
    hols = holidays.get(loc, [])
    result = {}
    for m in months:
        bd = biz_days.get(loc, {}).get(m, 0)
        leave_d = 0
        for row in leave_rows[1:]:
            if len(row) < 8: continue
            emp, start, end, ltype, status = (
                row[0].strip(), row[3].strip(), row[4].strip(),
                row[5].strip(), row[7].strip().lower()
            )
            if emp.lower() != name.lower(): continue
            if status not in VALID_STATUS: continue
            if ltype.lower() == "compensation": continue
            leave_d += count_leave_days(start, end, m, hols)
        result[m] = round((bd - leave_d) * 8 * factor)
    return result


# ── SHEETS HELPERS ───────────────────────────────────────────────────────────

def read_sheet(service, sheet_id, tab):
    result = service.spreadsheets().values().get(
        spreadsheetId=sheet_id,
        range=tab
    ).execute()
    return result.get("values", [])


def read_allocation(service):
    return read_sheet(service, ALLOCATION_SHEET_ID, ALLOCATION_TAB)


# ── RECONCILE ────────────────────────────────────────────────────────────────

def col_letter(n):
    """0-based column index to letter (e.g. 30 → AE)"""
    s = ""
    n += 1
    while n:
        n, r = divmod(n - 1, 26)
        s = chr(65 + r) + s
    return s


def reconcile_and_update(service, team, alloc_rows, leave_rows, biz_days, holidays, months, roster):
    """
    roster: [{name, login, loc, factor?}]

    Shared resources (logins in SHARED_RESOURCES) are updated in ALL allocation rows
    where they appear, regardless of the team's alloc_key — their leave data lives in
    whichever team's leave_rows is passed here.
    """
    alloc_key = team["alloc_key"]
    changes = []
    value_data = []
    format_requests = []

    login_map = {p["login"].lower(): p for p in roster}

    for i, row in enumerate(alloc_rows):
        team_val = row[1] if len(row) > 1 else ""
        login = row[7].strip().lower() if len(row) > 7 else ""
        if login not in login_map:
            continue
        # Shared resources update ALL rows where they appear.
        # Regular resources only update rows matching this team's alloc_key.
        if login not in SHARED_RESOURCES and alloc_key not in team_val:
            continue

        person = login_map[login]
        row_num = i + 1  # 1-based
        factor = SHARED_RESOURCES.get(login, person.get("factor", 1.0))

        hours_by_month = calc_person_hours(
            person["name"], person["loc"], leave_rows,
            biz_days, holidays, months, factor
        )

        for m in months:
            col_idx = MONTH_COL[m]
            current = int(row[col_idx]) if len(row) > col_idx and str(row[col_idx]).lstrip("-").isdigit() else 0
            calc_val = hours_by_month[m]
            if current == calc_val:
                continue

            cell_range = f"'{ALLOCATION_TAB}'!{col_letter(col_idx)}{row_num}"
            value_data.append({
                "range": cell_range,
                "values": [[calc_val]]
            })
            format_requests.append({
                "repeatCell": {
                    "range": {
                        "sheetId": ALLOCATION_SHEET_NUM,
                        "startRowIndex": i,
                        "endRowIndex": i + 1,
                        "startColumnIndex": col_idx,
                        "endColumnIndex": col_idx + 1,
                    },
                    "cell": {
                        "userEnteredFormat": {
                            "textFormat": {"bold": True, "foregroundColor": ORANGE}
                        }
                    },
                    "fields": "userEnteredFormat(textFormat)",
                }
            })
            changes.append({
                "team": team["name"], "login": login,
                "row": row_num, "month": m,
                "col": col_letter(col_idx),
                "old": current, "new": calc_val,
            })

    if value_data:
        service.spreadsheets().values().batchUpdate(
            spreadsheetId=ALLOCATION_SHEET_ID,
            body={"valueInputOption": "RAW", "data": value_data}
        ).execute()

    if format_requests:
        service.spreadsheets().batchUpdate(
            spreadsheetId=ALLOCATION_SHEET_ID,
            body={"requests": format_requests}
        ).execute()

    return changes


# ── ROSTER BUILDER ────────────────────────────────────────────────────────────

def build_roster_from_leave(leave_rows, alloc_rows, alloc_key):
    """Auto-build roster by cross-referencing leave sheet names with allocation logins.
    Requires the leave sheet to have a Location column (col L). If not available,
    set an explicit 'roster' in the team config instead."""
    team_logins = {}
    for row in alloc_rows:
        if len(row) > 7 and alloc_key in (row[1] if len(row) > 1 else ""):
            login = row[7].strip().lower()
            if login and login not in team_logins:
                team_logins[login] = True

    leave_people = {}
    for row in leave_rows[1:]:
        if not row or len(row) < 8:
            continue
        name = row[0].strip()
        loc_code = row[11].strip().upper() if len(row) > 11 and row[11] else ""
        loc = LOC_MAP.get(loc_code, None)
        if name and name not in leave_people and loc:
            leave_people[name] = loc

    roster = []
    matched_logins = set()

    for name, loc in leave_people.items():
        name_slug = name.lower().replace(" ", "").replace(".", "")
        best_login = None
        for login in team_logins:
            login_slug = login.replace(".", "").replace(" ", "")
            if login_slug in name_slug or name_slug.startswith(login_slug[:4]):
                best_login = login
                break
        if best_login and best_login not in matched_logins:
            matched_logins.add(best_login)
            roster.append({
                "name": name, "login": best_login, "loc": loc, "factor": 1.0
            })

    return roster


# ── GCHAT ────────────────────────────────────────────────────────────────────

def send_gchat(webhook_url, all_changes, run_date):
    if not all_changes:
        msg = f"✅ *Finance Hours — {run_date}*\nNenhuma divergência encontrada em todos os times."
    else:
        lines = [f"📊 *Finance Hours Reconciliation — {run_date}*\n"]
        by_team = {}
        for c in all_changes:
            by_team.setdefault(c["team"], []).append(c)

        month_names = {1:"Jan",2:"Feb",3:"Mar",4:"Apr",5:"May",6:"Jun",
                       7:"Jul",8:"Aug",9:"Sep",10:"Oct",11:"Nov",12:"Dec"}

        for team, changes in by_team.items():
            lines.append(f"*{team}* — {len(changes)} atualização(ões)")
            for c in changes:
                lines.append(
                    f"  • `{c['login']}` {month_names[c['month']]}: "
                    f"{c['old']}h → *{c['new']}h* (linha {c['row']})"
                )
            lines.append("")

        lines.append(f"_Total: {len(all_changes)} células atualizadas_")
        msg = "\n".join(lines)

    payload = {"text": msg}
    resp = requests.post(webhook_url, json=payload)
    resp.raise_for_status()


# ── MAIN ─────────────────────────────────────────────────────────────────────

def main():
    run_date = date.today().strftime("%d/%m/%Y")
    print(f"[{run_date}] Starting finance hours reconciliation...")

    today = date.today()
    months = list(range(today.month, 13))

    service      = get_sheets_service()
    github_token = get_github_token()
    webhook_url  = get_gchat_webhook()

    print("Loading business days from GitHub...")
    wb = download_biz_xlsx(github_token)
    biz_days, holidays = parse_biz_data(wb)

    print("Loading allocation sheet...")
    alloc_rows = read_allocation(service)

    all_changes = []

    for team in TEAMS:
        if team.get("skip"):
            print(f"  Skipping {team['name']} (no login mapping configured — see TODO in config).")
            continue

        print(f"Processing {team['name']}...")
        leave_rows = read_sheet(service, team["leave_sheet"], team["leave_tab"])

        if "roster" in team:
            roster = team["roster"]
            print(f"  Using explicit roster ({len(roster)} people).")
        else:
            roster = build_roster_from_leave(leave_rows, alloc_rows, team["alloc_key"])

        if not roster:
            print(f"  No roster found for {team['name']}, skipping.")
            continue

        changes = reconcile_and_update(
            service, team, alloc_rows, leave_rows,
            biz_days, holidays, months, roster
        )
        print(f"  {len(changes)} updates applied.")
        all_changes.extend(changes)

    print("Sending GChat notification...")
    send_gchat(webhook_url, all_changes, run_date)
    print(f"Done. {len(all_changes)} total updates.")


if __name__ == "__main__":
    main()
